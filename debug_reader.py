"""Debug: show ALL columns of the first rows of Marcaciones Nube.xlsx"""
import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Home\Downloads\Marcaciones Nube.xlsx", data_only=True)
ws = wb[wb.sheetnames[0]]

print(f"Sheet: {wb.sheetnames[0]}, Rows={ws.max_row}, Cols={ws.max_column}")
print(f"Merged: {[str(mc) for mc in ws.merged_cells.ranges]}")
print()

for row_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=16), start=1):
    non_empty = [(ci, v) for ci, v in enumerate(row) if v is not None and str(v).strip() != ""]
    if non_empty:
        # Print ALL columns, no truncation
        for ci, v in non_empty:
            print(f"  Row {row_idx:2d} C{ci:2d}: {repr(v)}")
        print()
    else:
        print(f"  Row {row_idx:2d}: (empty)\n")

wb.close()
