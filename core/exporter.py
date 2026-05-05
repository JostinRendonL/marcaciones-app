"""
core/exporter.py
Generates the formatted .xlsx attendance report using openpyxl.
"""
import io
import calendar
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.parser import parse_times, get_column_schema, assign_marks_to_columns, get_entry_exit_pairs
from core.reader import month_short


# ─────────────────────────────────────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────────────────────────────────────

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_GRAY_FILL    = PatternFill("solid", fgColor="F2F2F2")   # weekend rows
_HEADER_FILL  = PatternFill("solid", fgColor="D9E1F2")   # blue-ish header
_NAME_FILL    = PatternFill("solid", fgColor="C00000")   # red name banner
_PAGAR_FILL   = PatternFill("solid", fgColor="E2EFDA")   # light green pay header
_WARN_FILL    = PatternFill("solid", fgColor="FF6600")   # orange warning cell
_PAGAR_H_FILL = PatternFill("solid", fgColor="70AD47")   # dark-green pay header

# Pay rates (USD / hour)
_RATE_WEEKDAY = 3.26
_RATE_WEEKEND = 6.27

_DAYS_ES = {
    0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves",
    4: "Viernes", 5: "Sábado", 6: "Domingo",
}


def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col)


def _write(ws, row: int, col: int, value,
           bold: bool = False, fill=None,
           color: str = "000000", size: int = 10,
           h_align: str = "center"):
    c = ws.cell(row=row, column=col)
    c.value = str(value) if value is not None and value != "" else ""
    c.font = Font(bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    c.border = _BORDER
    c.number_format = "@"          # Force text — no Excel autoformat
    if fill:
        c.fill = fill
    return c


# ─────────────────────────────────────────────────────────────────────────────
# Sheet name builder
# ─────────────────────────────────────────────────────────────────────────────

def _sanitize_sheet_name(name: str) -> str:
    """Remove characters invalid in Excel sheet names: \\ / ? * [ ] :"""
    for ch in ("\\", "/", "?", "*", "[", "]", ":"):
        name = name.replace(ch, "")
    # Also strip leading/trailing apostrophes (Excel quirk)
    name = name.strip().strip("'")
    return name if name else "SinNombre"


def build_sheet_name(name: str, month: int, year: int) -> str:
    """
    Format: "{Name} {MonthShort}{Year2d}"  e.g. "VICKY Jul25"
    Sanitized and truncated to 31 chars (Excel limit).
    """
    yr2 = str(year)[-2:]
    sheet = f"{name} {month_short(month)}{yr2}"
    sheet = _sanitize_sheet_name(sheet)
    return sheet[:31]


# ─────────────────────────────────────────────────────────────────────────────
# Single-person sheet writer
# ─────────────────────────────────────────────────────────────────────────────

def _write_person_sheet(wb, record: dict, edited_days: Optional[dict] = None):
    """
    Write one worksheet for a person+month record.

    record = { name, month, year, days: {day_int: raw_string} }
    edited_days: optional overrides { day_int: raw_string } from user corrections
    """
    name  = record["name"]
    month = record["month"] or 1
    year  = record["year"] or 2025
    days  = dict(record["days"])  # copy

    if edited_days:
        days.update(edited_days)

    # ── Determine column schema ────────────────────────────────────────────
    all_parsed = {d: parse_times(raw, name, str(d)) for d, raw in days.items()}
    max_marks  = max((len(v) for v in all_parsed.values()), default=0)
    schema     = get_column_schema(max_marks)

    # ── Create / get sheet ────────────────────────────────────────────────
    sheet_name = build_sheet_name(name, month, year)
    # Handle duplicate names
    base = sheet_name
    suffix = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base[:29]}{suffix}"
        suffix += 1

    ws = wb.create_sheet(title=sheet_name)

    # ── Row 1: Person name (red, bold, merged) ────────────────────────────
    total_cols = 2 + len(schema)  # FECHA + DIA + mark cols
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    name_cell = ws.cell(row=1, column=1)
    name_cell.value = name
    name_cell.font  = Font(bold=True, size=12, color="FFFFFF")
    name_cell.fill  = PatternFill("solid", fgColor="C00000")
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.number_format = "@"

    # ── Row 2: Período ────────────────────────────────────────────────────
    days_in_month = calendar.monthrange(year, month)[1]
    period_str = (
        f"Período: 01/{month:02d}/{year} ~ {days_in_month:02d}/{month:02d}/{year}"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    p_cell = ws.cell(row=2, column=1)
    p_cell.value = period_str
    p_cell.font  = Font(bold=False, size=10)
    p_cell.alignment = Alignment(horizontal="center", vertical="center")
    p_cell.number_format = "@"

    # ── Row 3: blank ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Headers ───────────────────────────────────────────────────
    headers = ["FECHA", "DIA"] + schema + ["PAGAR"]
    for ci, h in enumerate(headers, start=1):
        if h == "PAGAR":
            _write(ws, 4, ci, h, bold=True, fill=_PAGAR_H_FILL,
                   color="FFFFFF", size=10)
        else:
            _write(ws, 4, ci, h, bold=True, fill=_HEADER_FILL, size=10)

    # Pre-compute (entry_col, exit_col) pairs (0-based within schema)
    pairs = get_entry_exit_pairs(schema)
    # Column numbers in the sheet (1-based): FECHA=1, DIA=2, schema starts at 3
    schema_offset = 3  # first schema column
    pagar_col = schema_offset + len(schema)  # last column

    # ── Row 5+: Data rows ─────────────────────────────────────────────────
    for day_num in range(1, days_in_month + 1):
        data_row = 4 + day_num

        try:
            day_date = date(year, month, day_num)
            weekday  = day_date.weekday()  # 0=Mon … 6=Sun
        except ValueError:
            continue

        fecha_str = day_date.strftime("%d/%m/%Y")
        dia_str   = _DAYS_ES[weekday]

        is_weekend = weekday >= 5
        row_fill = _GRAY_FILL if is_weekend else None
        rate     = _RATE_WEEKEND if is_weekend else _RATE_WEEKDAY

        # Write FECHA and DIA
        _write(ws, data_row, 1, fecha_str, fill=row_fill)
        _write(ws, data_row, 2, dia_str,   fill=row_fill)

        # Parse and assign
        raw    = days.get(day_num, "")
        parsed = parse_times(raw, name, str(day_num))
        mapped = assign_marks_to_columns(parsed, schema)
        pos    = mapped.get("_pos", {})

        # Write time cells using positional values (handles duplicate col names)
        for ci, col_name in enumerate(schema, start=schema_offset):
            pos_idx = ci - schema_offset  # 0-based
            val = pos.get(pos_idx, mapped.get(col_name, "")) if pos else mapped.get(col_name, "")
            _write(ws, data_row, ci, val, fill=row_fill)

        # ── PAGAR calculation ─────────────────────────────────────────────
        if not pairs or not parsed:
            # No time at all → leave blank
            _write(ws, data_row, pagar_col, "", fill=row_fill)
        else:
            total_hours = 0.0
            has_warning = False
            for (entry_idx, exit_idx) in pairs:
                entry_val = pos.get(entry_idx, "") if pos else ""
                exit_val  = pos.get(exit_idx,  "") if pos else ""

                # If exactly one of the two is missing → mark as warning
                if bool(entry_val) != bool(exit_val):
                    has_warning = True
                    break

                # Both present → accumulate hours
                if entry_val and exit_val:
                    try:
                        eh, em = map(int, entry_val.split(":"))
                        xh, xm = map(int, exit_val.split(":"))
                        delta_h = (xh * 60 + xm - eh * 60 - em) / 60.0
                        if delta_h < 0:
                            delta_h += 24  # past midnight
                        total_hours += delta_h
                    except (ValueError, TypeError):
                        has_warning = True
                        break

            if has_warning:
                c = ws.cell(row=data_row, column=pagar_col)
                c.value = "⚠ Revisar"
                c.font  = Font(bold=True, size=10, color="FFFFFF")
                c.fill  = _WARN_FILL
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = _BORDER
                c.number_format = "@"
            else:
                pay = round(total_hours * rate, 2)
                _write(ws, data_row, pagar_col,
                       f"${pay:.2f}" if pay else "",
                       fill=row_fill)

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = {"FECHA": 13, "DIA": 11}
    for col_name in schema:
        col_widths[col_name] = 13

    all_headers = ["FECHA", "DIA"] + schema + ["PAGAR"]
    widths_list = [col_widths.get(h, 13) for h in all_headers]
    widths_list[-1] = 13  # PAGAR column

    for ci, w in enumerate(widths_list, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[4].height = 18

    # ── Freeze panes so header stays visible ──────────────────────────────
    ws.freeze_panes = "C5"  # freeze FECHA + DIA columns and first 4 rows


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def export_to_xlsx(records: list[dict], edited_overrides: Optional[dict] = None) -> bytes:
    """
    Generate the full report workbook.

    records: list of person records from reader.py
    edited_overrides: { (name, month, year, day): raw_string } corrections from UI

    Returns bytes of the .xlsx file.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for rec in records:
        # Build per-record day overrides
        person_edits = {}
        if edited_overrides:
            for (n, m, y, d), raw in edited_overrides.items():
                if n == rec["name"] and m == rec["month"] and y == rec["year"]:
                    person_edits[d] = raw

        try:
            _write_person_sheet(wb, rec, person_edits if person_edits else None)
        except Exception as exc:
            # Add an error sheet instead of crashing
            err_sheet_name = _sanitize_sheet_name(f"ERROR_{rec.get('name','?')[:20]}")
            err_sheet_name = err_sheet_name[:31]
            ws = wb.create_sheet(title=err_sheet_name)
            ws.cell(row=1, column=1).value = f"Error procesando: {exc}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
