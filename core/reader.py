"""
core/reader.py
Handles reading and normalizing biometric .xlsx files.

Supports THREE input formats (auto-detected per sheet):

  A) Pre-processed table (one sheet per person, already clean):
       Row 1: "{Name} – {Month} {Year}"  (merged)
       Row 2: "Período: yyyy-mm-dd ~ yyyy-mm-dd"  (merged)
       Row 4: Headers — FECHA | DIA | INGRESO | SALIDA | ...
       Row 5+: Data rows with dates and times in columns

  B) Raw biometric matrix (Marcaciones Nube format):
       Row:  Periodo: | ... | '2025-07-01 ~ 2025-07-31'
       Row:  1 | 2 | 3 | ... | 31                          ← day-number header
       Row:  ID: | ... | cédula | ... | Nombre: | ... | name
       Row:  '09:5912:3319:13' | ... | glued times          ← data per day
       (repeat ID+data rows for each person)
       (empty row)
       (next Periodo: block)

  C) Raw biometric WITHOUT period row (marcacioness format):
       Row:  1 | 2 | 3 | ... | 31                          ← starts directly
       Row:  ID: | ... | name
       Row:  data
       ...
"""
import re
import io
from datetime import date, datetime
from typing import Optional, Tuple, List, Dict

import pandas as pd
import openpyxl


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

_MONTH_MAP = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "ene": 1, "feb": 2, "mar": 3, "abr": 4,
    "jun": 6, "jul": 7, "ago": 8,
    "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}

_MONTH_SHORT_ES = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Ago",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}


def month_short(month_num: int) -> str:
    return _MONTH_SHORT_ES.get(month_num, str(month_num))


def parse_period_from_text(text: str) -> Tuple[Optional[int], Optional[int]]:
    """
    Extract (month, year) from a period-style cell value.
    Handles:
      - "2025-07-01 ~ 2025-07-31"
      - "Período: 2025-07-01 ~ 2025-07-31"
      - "01/07/2025 ~ 31/07/2025"
      - "julio 2025"
    """
    text = str(text).strip()

    # yyyy-mm-dd pattern
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if m:
        return int(m.group(2)), int(m.group(1))

    # dd/mm/yyyy pattern
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if m:
        return int(m.group(2)), int(m.group(3))

    # Month name + year
    text_lower = text.lower()
    for name, num in _MONTH_MAP.items():
        if name in text_lower:
            yr = re.search(r"(\d{4}|\d{2})", text)
            if yr:
                y = int(yr.group(1))
                if y < 100:
                    y += 2000
                return num, y
    return None, None


def _parse_date_cell(val) -> Optional[date]:
    """Parse a date from a cell value (datetime, date, or string)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    # dd/mm/yyyy
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    # yyyy-mm-dd
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None


def _is_name_text(s: str) -> bool:
    """Return True if the string looks like a person name (no colons, not purely numeric)."""
    s = s.strip()
    return bool(s) and ":" not in s and not s.replace(" ", "").replace("_", "").replace("-", "").isdigit()


def _extract_person_name(row) -> Optional[str]:
    """
    Extract person name from an ID row.

    Handles two formats:
      Standard:  C0='ID:' C2=cédula/name C10=name
      Alternate: C0=empty, C2=name (no 'ID:' prefix — some biometric exports)
    """
    cells = list(row)
    first = str(cells[0] or "").strip().upper() if cells else ""

    # ── Standard format: starts with "ID:" ────────────────────────────────
    if first == "ID:":
        name = ""
        # Try C10 first (Nombre: field)
        if len(cells) > 10 and cells[10] is not None:
            name = str(cells[10]).strip()

        # Try C8 — some exports store "Nombre: [name]" or "Nombre:\n[name]" there
        if (not name or len(name) <= 1) and len(cells) > 8 and cells[8] is not None:
            col8 = str(cells[8]).strip()
            if re.search(r'(?i)nombre\s*:', col8):
                extracted = re.sub(r'(?i)^.*nombre\s*:\s*\n?\s*', '', col8).strip()
                if extracted and len(extracted) > 1 and not extracted.replace(" ", "").isdigit():
                    name = extracted

        # Fall back to C2 if still empty (name may be in the cédula column)
        if (not name or len(name) <= 1) and len(cells) > 2 and cells[2] is not None:
            candidate = str(cells[2]).strip()
            if not candidate.replace(" ", "").isdigit():
                name = candidate

        # Last resort: C20 (Departamento value sometimes holds first name)
        if not name and len(cells) > 20 and cells[20] is not None:
            name = str(cells[20]).strip()

        return name.strip() if name else None

    # ── Alternate format: name directly in C2, C0 empty ───────────────────
    if first == "" and len(cells) > 2:
        candidate = str(cells[2] or "").strip()
        if _is_name_text(candidate):
            return candidate

    return None


# ─────────────────────────────────────────────────────────────────────────────
# Format detection
# ─────────────────────────────────────────────────────────────────────────────

def _is_preprocessed_sheet(ws) -> bool:
    """
    A sheet is pre-processed if row 4 contains headers like FECHA, DIA, INGRESO.
    """
    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    if not row4:
        return False
    cells = [str(c or "").strip().upper() for c in row4[0]]
    return "FECHA" in cells and "DIA" in cells


def _is_day_number_row(row) -> bool:
    """Check if a row contains sequential day numbers (1, 2, 3... up to 28-31)."""
    nums = []
    for c in row:
        if c is None:
            continue
        try:
            n = int(float(str(c).strip()))
            if 1 <= n <= 31:
                nums.append(n)
        except (ValueError, TypeError):
            pass
    return len(nums) >= 10  # At least 10 day numbers


def _is_period_row(row) -> bool:
    """Check if row starts with 'Periodo:' or 'Período:'."""
    first = str(row[0] or "").strip().lower() if row else ""
    return "periodo" in first or "período" in first


def _is_id_row(row) -> bool:
    """Check if row starts with 'ID:' or has a name in col[2] with col[0] empty."""
    if not row:
        return False
    first = str(row[0] or "").strip().upper()
    if first == "ID:":
        return True
    # Alternate format: name in col[2], col[0] empty (some biometric exports omit 'ID:')
    if first == "" and len(row) > 2:
        candidate = str(row[2] or "").strip()
        if _is_name_text(candidate):
            return True
    return False


# ─────────────────────────────────────────────────────────────────────────────
# Format A: Pre-processed table (one sheet per person)
# ─────────────────────────────────────────────────────────────────────────────

def _parse_name_from_title(text: str) -> str:
    """'Nizurieta – Julio 2025' → 'NIZURIETA'"""
    text = str(text).strip()
    for sep in ["–", "—", "-", "·"]:
        if sep in text:
            candidate = text.split(sep, 1)[0].strip()
            if candidate:
                return candidate.upper()
    return re.sub(r"\d+$", "", text).strip().upper()


def _read_preprocessed_sheet(ws, sheet_name: str) -> Optional[dict]:
    """Read a pre-processed sheet and return a record dict."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        return None

    title = str(rows[0][0] or "").strip()
    name = _parse_name_from_title(title)

    period_text = str(rows[1][0] or "").strip()
    month, year = parse_period_from_text(period_text)
    if month is None:
        month, year = parse_period_from_text(title)
    if month is None:
        month, year = parse_period_from_text(sheet_name)

    # Row 4: Headers
    header_row = rows[3] if len(rows) > 3 else []
    headers = [str(c or "").strip().upper() for c in header_row]
    fecha_idx = headers.index("FECHA") if "FECHA" in headers else 0

    # Time columns
    time_cols = []
    for ci, h in enumerate(headers):
        if h and h not in ("FECHA", "DIA", ""):
            time_cols.append((ci, h))

    # Data rows
    days_dict = {}
    for row in rows[4:]:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        fecha_val = row[fecha_idx] if fecha_idx < len(row) else None
        day_date = _parse_date_cell(fecha_val)
        if day_date is None:
            continue

        times = []
        for ci, _ in time_cols:
            if ci < len(row):
                val = row[ci]
                if val is not None and str(val).strip() != "":
                    times.append(str(val).strip())

        if times:
            days_dict[day_date.day] = " ".join(times)

    if month is None and days_dict:
        for row in rows[4:]:
            d = _parse_date_cell(row[fecha_idx] if fecha_idx < len(row) else None)
            if d:
                month, year = d.month, d.year
                break

    if not name:
        name = sheet_name.upper()

    return {
        "name": name,
        "month": month or 1,
        "year": year or 2025,
        "days": days_dict,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Format B/C: Raw biometric matrix
# ─────────────────────────────────────────────────────────────────────────────

def _read_raw_biometric_sheet(ws) -> List[dict]:
    """
    Parse a raw biometric sheet.

    Structure per block:
      [optional] Periodo: row
      Day-number header: 1, 2, 3, ... 31  (col index = day - 1 IF starting at C0)
      Pairs of (ID row, data row) for each person
      Empty row → next block
    """
    rows = list(ws.iter_rows(values_only=True))
    records = []

    current_month: Optional[int] = None
    current_year: Optional[int] = None
    col_to_day: Dict[int, int] = {}  # column-index → day-number

    i = 0
    while i < len(rows):
        row = rows[i]

        # Skip fully empty rows
        if all(c is None or str(c).strip() == "" for c in row):
            i += 1
            continue

        # ── Period row ──────────────────────────────────────────────────
        if _is_period_row(row):
            # Period date range is usually at C2
            for ci in range(len(row)):
                val = row[ci]
                if val is None:
                    continue
                s = str(val).strip()
                if "~" in s or re.search(r"\d{4}-\d{2}-\d{2}", s):
                    m, y = parse_period_from_text(s)
                    if m and y:
                        current_month, current_year = m, y
                        break
            # Some exports have the correct end-date in a later column (e.g. col[11])
            # while col[2] still shows the previous month — use the later date if newer.
            for ci in range(len(row)):
                val = row[ci]
                if val is None:
                    continue
                s = str(val).strip()
                if re.search(r"\d{4}-\d{2}-\d{2}", s) and "~" not in s:
                    m2, y2 = parse_period_from_text(s)
                    if m2 and y2:
                        if current_year is None or y2 > current_year or (y2 == current_year and m2 > current_month):
                            current_month, current_year = m2, y2
                        break
            i += 1
            continue

        # ── Day-number header row ───────────────────────────────────────
        if _is_day_number_row(row):
            col_to_day = {}
            for ci, val in enumerate(row):
                if val is None:
                    continue
                try:
                    n = int(float(str(val).strip()))
                    if 1 <= n <= 31:
                        col_to_day[ci] = n
                except (ValueError, TypeError):
                    pass
            i += 1
            continue

        # ── ID row → read person + data ─────────────────────────────────
        if _is_id_row(row) and col_to_day:
            name = _extract_person_name(row)
            if not name:
                name = "DESCONOCIDO"

            # Next row should be the data row
            data_row_idx = i + 1
            if data_row_idx < len(rows):
                data_row = rows[data_row_idx]
                days_dict = {}
                for ci, day_num in col_to_day.items():
                    if ci < len(data_row):
                        val = data_row[ci]
                        if val is not None and str(val).strip() not in ("", "nan"):
                            days_dict[day_num] = str(val).strip()

                records.append({
                    "name": name.upper(),
                    "month": current_month,
                    "year": current_year,
                    "days": days_dict,
                })
                i = data_row_idx + 1
            else:
                i += 1
            continue

        # ── Unknown row → skip ──────────────────────────────────────────
        i += 1

    return records


# ─────────────────────────────────────────────────────────────────────────────
# Main public entry point
# ─────────────────────────────────────────────────────────────────────────────

def read_biometric_xlsx(file_bytes: bytes) -> List[dict]:
    """
    Parse a biometric .xlsx file.
    Auto-detects whether sheets are pre-processed (Format A) or raw (Format B/C).

    Returns a list of dicts, one per (person, period):
        { "name": str, "month": int, "year": int, "days": {day_int: raw_string} }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        try:
            if _is_preprocessed_sheet(ws):
                rec = _read_preprocessed_sheet(ws, sheet_name)
                if rec and rec["days"]:
                    records.append(rec)
            else:
                sheet_records = _read_raw_biometric_sheet(ws)
                records.extend(sheet_records)
        except Exception as exc:
            records.append({
                "name": f"ERROR {sheet_name}",
                "month": 1,
                "year": 2025,
                "days": {},
                "_error": str(exc),
            })

    wb.close()
    return records


def records_to_dataframe(records: List[dict]) -> pd.DataFrame:
    """
    Flatten the raw records list into a tidy DataFrame for preview/editing.
    """
    rows = []
    for rec in records:
        for day, raw in sorted(rec["days"].items()):
            rows.append({
                "Persona": rec["name"],
                "Mes": rec["month"],
                "Año": rec["year"],
                "Día": int(day),
                "Marcaciones Raw": raw,
            })
    if not rows:
        return pd.DataFrame(columns=["Persona", "Mes", "Año", "Día", "Marcaciones Raw"])
    return pd.DataFrame(rows)
